from win32api import GetFileVersionInfo, LOWORD, HIWORD
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
import os
import getpass
import platform


explorer = 0
alzip = 0
flash = 0
hancom= 0
kakao=0


explorer_url="https://support.microsoft.com/ko-kr/help/17621/internet-explorer-downloads"
alzip_url="알집 자동 업데이트 혹은 https://www.altools.co.kr/Download/ALZip.aspx 에서 새로 설치"
flash_url="https://get.adobe.com/kr/flashplayer/"
hancom_url="https://www.hancom.com/cs_center/csDownload.do?utm_source=download.hancom.com&utm_medium=URL_Redirect&utm_campaign=Redirect"
kakao_url="PC 카카오톡 - 설정 - 카카오톡 정보 - 업데이트"

for (path, dir, files) in os.walk("C:\\") or os.walk("D:\\"):
    for fn in files:
        ext = os.path.basename(fn)
        if ext == "iexplore.exe" and explorer == 0:
            explorer = path + '\\' + fn
            break
        if ext == "ALZip.exe" and alzip == 0:
            alzip = path + '\\' + fn
            break
        if ext == "FlashUtil64_32_0_0_223_ActiveX.exe" and flash ==0:
            flash = path + '\\' + fn
            break
        if ext == "HancomStudio.exe" and hancom == 0:
            hancom = path + '\\' + fn
            break
        if ext == "KakaoTalk.exe" and kakao == 0:
            kakao = path + '\\' + fn
            break
    if explorer and alzip and flash and hancom and kakao is not None:
        break


def get_version_number(filename):
    info = GetFileVersionInfo(filename, "\\")
    ms = info['FileVersionMS']
    ls = info['FileVersionLS']
    return HIWORD(ms), LOWORD(ms), HIWORD(ls), LOWORD(ls)


write_wb = Workbook()
write_ws = write_wb.active
file_name = ('C:\\Users\\' + getpass.getuser() + '\Desktop\Result.xlsx')
write_ws.title = "Exploit List"

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
highlight.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# line = NamedStyle(name="line")
# bd = Side(style='thick', color="000000")
# line.border = Border(left=bd, top=bd, right=bd, bottom=bd)

def check(file, version):
    if 'iexplore'in file and version <= '10':
        return "취약한 버전입니다! 최신버전으로 업데이트 바랍니다."
    if 'Flash' in file and version <= '10':
        return "취약한 버전입니다! 최신버전으로 업데이트 바랍니다."
    if 'alzip' in file and version <= '10':
        return "취약한 버전입니다! 최신버전으로 업데이트 바랍니다."
    if 'Hancom' in file and version <= '10.0.0.8':
        return "취약한 버전입니다! 최신버전으로 업데이트 바랍니다."
    if 'Kakao' in file and version <= '2.7':
        return "취약한 버전입니다! 최신버전으로 업데이트 바랍니다."
    else:
        return "안전한 버전입니다!"



write_ws.column_dimensions["A"].width = 20
write_ws.column_dimensions["B"].width = 20
write_ws.column_dimensions["C"].width = 20
write_ws.column_dimensions["D"].width = 40

write_ws.append(["OS"]+(["OS Version"])+(["OS Version Check"]))
write_ws.append([platform.platform()]+([platform.version()])+(["안전한 버전입니다."]))
write_ws.append([""]+([""])+([""]))
write_ws.append(["파일명"]+(["파일의 현재 버전"])+(["Version Check"])+(["대응 방안"]))
if (explorer != 0):
    explorer_chk = ".".join([str(i) for i in get_version_number(explorer)])
    write_ws.append(["Explorer"] + ([explorer_chk]) + ([check(explorer,explorer_chk)]) + ([explorer_url]))
if (alzip != 0):
    alzip_chk = ".".join([str(i) for i in get_version_number(alzip)])
    write_ws.append(["ALZip"] + ([alzip_chk]) + ([check(alzip,alzip_chk)])+ ([alzip_url]))
if (flash != 0):
    flash_chk = ".".join([str(i) for i in get_version_number(flash)])
    write_ws.append(["Flash Player"] + ([flash_chk]) + ([check(flash,flash_chk)]) + ([flash_url]))
if (hancom != 0):
    hancom_chk = ".".join([str(i) for i in get_version_number(hancom)])
    write_ws.append(["Hancom Office"] + ([hancom_chk]) + ([check(hancom,hancom_chk)])+ ([hancom_url]))
if (kakao != 0):
    kakao_chk = ".".join([str(i) for i in get_version_number(kakao)])
    write_ws.append(["KakaoTalk"] + ([kakao_chk]) + ([check(kakao,kakao_chk)])+ ([kakao_url]))

rows = range(1, 44)
columns = range(1, 10)
for row in rows:
    for col in columns:
        if row == 1:
            write_ws.cell(row, col).style = highlight
        if row == 4:
            write_ws.cell(row, col).style = highlight
        write_ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


write_wb.save(filename=file_name)
